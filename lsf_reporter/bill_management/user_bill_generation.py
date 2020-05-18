# -*- coding: utf-8 -*-


import os
import math
import shutil
import logging
from dateutil import rrule
from datetime import datetime
import calendar
import json
import openpyxl
from collections import OrderedDict
from pprint import pprint
from elasticsearch import Elasticsearch, helpers
from lsf_reporter.bill_management.time_counter import TimeCounter
from lsf_reporter.elasticsearch_client import init_conn


class CustomJobBilling():
    """
    自定义查询并生成  计费账单
    """
    def __init__(self, user_name, queue_name_list, start_date, end_date, storage_list, consul_queue_dict, es_url):
        """
        初始值
        """
        self.storage_list = storage_list
        self.consul_queue_dict = consul_queue_dict
        self.ES = Elasticsearch(hosts=es_url)

        self.user_name = user_name
        self.index_name = self.storage_list[0]
        self.queue_name_list = queue_name_list

        self.start_date = datetime.strptime(start_date, "%Y-%m-%d")
        self.end_date = datetime.strptime(end_date, "%Y-%m-%d")
        self.year = self.start_date.year
        self.month = self.start_date.month
        self.day = self.start_date.day
        self.timespan = rrule.rrule(freq=rrule.DAILY, dtstart=self.start_date, until=self.end_date).count() - 1

        self.templatePath = self.storage_list[4]
        self.bill_name_for_admin = "{0}_{1}_{2}-{3}".format(self.index_name, self.user_name, self.year, self.month)
        self.billPath_for_admin = "{0}/{1}/{2}/{3}.xlsx".format(self.storage_list[2], self.year, self.month, self.bill_name_for_admin)
        self.bill_name_for_user = "{}_{}_from_{}_to_{}".format(self.index_name, self.user_name, start_date, end_date)
        self.billPath_for_user = "{0}/{1}/{2}.xlsx".format(self.storage_list[3], self.user_name, self.bill_name_for_user)
        # self.tmpStorePath = "/dev/shm/lsf_reporter/{0}".format(self.user_name)  # for /dev/shm in linux
        self.tmpStorePath = self.storage_list[3]  # for windows test
        self.tmpbillPath_for_admin_and_or_user = "{}/{}.xlsx".format(self.tmpStorePath, self.bill_name_for_user)


    def time_consumption(self, t):
        """
        转换时间戳
        """
        timeArray = datetime.strptime(t, "%Y-%m-%dT%H:%M:%S.000Z")
        return timeArray


    def query_queue(self, user_name):
        """
        通过es的聚合查询，获取某用户在一定时间，有提交记录的队列名列表
        """
        es = self.ES
        body = {
            "size": 0,  # 只返回聚合结果，不返回查询数据
            "query": {
                "bool": {
                    "filter": {
                        "range": {
                            "@timestamp": {
                                "from": self.start_date,
                                "to": self.end_date
                            }
                        }
                    },
                    "must": [
                        {"term": {
                            "user_name": {
                                "value": user_name  # 查询用户
                            }
                        }}
                    ]
                }
            },
            "aggs": {
                "queue_name": {
                    "terms": {
                        "field": "queue_name"  # 查询——队列——的聚合结果
                    }
                }
            }
        }
        query_queue = es.search(index=self.index_name, body=body)
        query_queue = query_queue["aggregations"]["queue_name"]["buckets"]
        query_queue_name_list = []
        query_queue_name_record_list = []
        if len(query_queue) > 0:
            query_queue_name_list = sorted([item['key'] for item in query_queue])
            query_queue_name_record_list = [{item['key']: item['doc_count']} for item in query_queue]
            logging.info("检索到 {0} 用户在 {1} 日志中，有记录的提交队列列表为：{2}".format(user_name, self.index_name, query_queue_name_record_list))
        elif len(query_queue) == 0:
            logging.warning("未检测到用户 {0} 在 {1} 日志中的作业提交记录".format(user_name, self.index_name))
            query_queue_name_list.append('Empty_Record')
            query_queue_name_record_list.append({'Empty_Record': 0})
        return query_queue_name_list  # 其中 queue_name_record_list 中包含 {队列名：相应日志条数} 的信息


    def query_total_log(self, queue_name_tmp):
        """
        定制的 DSL 请求语句，用户名、起始时间、终止时间、队列名 ———— 获取完整数据
        """
        es = self.ES
        body = {
            "_source": {
                "includes": ["user_name", "queue_name", "num_exec_hosts", "exec_hosts",
                             "job_id", "job_name", "job_status", "submit_time", "start_time", "end_time"]
            },
            "query": {
                "bool": {
                    "filter": {
                        "range": {
                            "@timestamp": {
                                "from": self.start_date,
                                "to": self.end_date
                            }
                        }
                    },
                    "must": [
                        {"match": {"user_name": self.user_name}},
                        {"match": {"queue_name": queue_name_tmp}}
                    ]
                }
            }
        }
        try:
            query_total_log = helpers.scan(client=es, query=body, scroll="5m", size=2000, clear_scroll=True)
            total_logs = [log for log in query_total_log]
            return total_logs
        except MemoryError as e:
            logging.warning(e)
            exit(1)


    def total_job_result(self, queue_name_tmp):
        """
        作业日志的格式转换，构造合适的json串
        """
        jobconsul = self.consul_queue_dict
        logsource = self.query_total_log(queue_name_tmp)
        joblogdata = [logsource[i]['_source'] for i in range(len(logsource))]

        job_result = []
        for i in range(len(joblogdata)):
            jobqueue_i = joblogdata[i].get('queue_name')  # 为了匹配 consul中的 队列名
            num_exec_hosts_i = joblogdata[i].get('num_exec_hosts')
            joblogdata[i]['submit_time'] = self.time_consumption(joblogdata[i].get('submit_time'))
            joblogdata[i]['start_time'] = self.time_consumption(joblogdata[i].get('start_time'))
            joblogdata[i]['end_time'] = self.time_consumption(joblogdata[i].get('end_time'))
            submit_time_i = joblogdata[i].get('submit_time')  # 后期可能修改时间值
            start_time_i = joblogdata[i].get('start_time')
            end_time_i = joblogdata[i].get('end_time')

            run_duration_i = 0.0  # 如果是日志记录的错误作业，则不计费
            if start_time_i >= submit_time_i:
                run_duration_i = int(math.ceil((end_time_i - start_time_i).total_seconds()))  # 按秒数运算
                joblogdata[i]["exec_hosts"] = str(joblogdata[i].get('exec_hosts'))  # 将  [{'num_slots': 64, 'host_name': 'node-arm02'}] 强制转换为str
            elif start_time_i < submit_time_i:
                joblogdata[i]["exec_hosts"] = str("[None]")  # 返回空值
            jobtime_i = run_duration_i * num_exec_hosts_i
            jobprice_i = float(jobconsul[jobqueue_i].get('Price'))
            jobcost_i = float(jobprice_i * jobtime_i / int(3600))
            # 构造新的 json
            keys = ['jobprice', 'jobcost', 'run_duration']
            values = [jobprice_i, jobcost_i, run_duration_i]
            jobtmp_i = dict(zip(keys, values))
            job_result_i = dict(list(joblogdata[i].items()) + list(jobtmp_i.items()))
            job_result.append(job_result_i)

        # 将数据有序排列
        ordered_job_result = []
        ORDERED_KEYS = ['job_id', 'exec_hosts', 'submit_time', 'start_time', 'end_time', 'job_status',
                        'run_duration', 'num_exec_hosts']
        for jr in job_result:
            ordered_jr = OrderedDict((k, jr[k]) for k in ORDERED_KEYS)
            ordered_job_result.append(ordered_jr)
        ordered_job_result.sort(key=lambda x: x["job_id"])  # 将作业按照 job_id 排序
        # pprint(ordered_job_result[0])  #  测试生成的作业账单数据
        return ordered_job_result


    def generate_bill_template(self, queue_name_list):
        """
        生成相应的原始账单模板——存放在 /dev/shm 上面
        """
        if not os.path.exists(self.storage_list[4]):  # 每次在doc/bill_template下更新模板，自动判断传到存储位置
            os.makedirs(os.path.dirname(self.storage_list[4]))
        shutil.copyfile("doc/bill_template/{0}".format(os.path.basename(self.storage_list[4])), self.storage_list[4])
        if not os.path.exists(self.tmpStorePath):
            os.makedirs(self.tmpStorePath)

        wb = openpyxl.load_workbook(self.templatePath)
        for q in range(len(queue_name_list)):
            copy_sheet_q = wb.copy_worksheet(wb.worksheets[1])  # 复制 detailed 表
            copy_sheet_q.title = queue_name_list[q]
        wb.remove(wb.worksheets[1])
        wb.save(self.tmpbillPath_for_admin_and_or_user)
        return self.tmpbillPath_for_admin_and_or_user


    def json_to_excel(self, queue_name_tmp):
        """
        按账单模板格式输出作业信息至 excel
        """
        if queue_name_tmp == 'Empty_Record':
            job_result = []
        else:
            # 将所有已经准备好的队列数据，逐个写入。
            logging.info("正在导出用户 {0} 在队列 {1} 的计费数据".format(self.user_name, queue_name_tmp))
            listHead = []
            ordered_job_result = self.total_job_result(queue_name_tmp)
            wb = openpyxl.load_workbook(self.tmpbillPath_for_admin_and_or_user)
            ws = wb[queue_name_tmp]
            # 填充一些基本账单信息
            ws.cell(row=2, column=2, value=self.user_name)  # 用户名
            ws.cell(row=3, column=2, value=queue_name_tmp)  # 队列名
            ws.cell(row=3, column=4, value=self.consul_queue_dict[queue_name_tmp].get('Price'))  # 队列计费单价
            ws.cell(row=4, column=2, value=self.start_date)  # 起始日
            ws.cell(row=4, column=4, value=self.end_date)  # 截止日
            ws.cell(row=4, column=6, value=self.timespan)  # 查询时间跨度
            ws.cell(row=5, column=6, value=int(len(ordered_job_result)))  # 作业数

            for c, i in enumerate(ordered_job_result[0].keys()):  # 写入json值的第一列 key值
                ws.cell(row=12, column=c + 1, value=i)
                listHead.append(i)  # listHead 是excel第一行的值

            for r, i in enumerate(ordered_job_result):
                row = r + 12  # 从账单的第十二行开始写
                for c, d in enumerate(listHead):  # column——data[d]
                    ws.cell(row=row, column=c + 1, value=i.get(d, ""))
            listHead.clear()
            for i in range(len(ordered_job_result) + 12, 20000):
                ws.cell(row=i, column=9, value="")  # 最后清空没有值的列
            wb.save(self.tmpbillPath_for_admin_and_or_user)


    def brief_bill_info_gather(self):
        wb = openpyxl.load_workbook(self.tmpbillPath_for_admin_and_or_user)
        ws_brief = wb['brief']
        ws_brief.cell(row=2, column=2, value=self.user_name)
        ws_brief.cell(row=2, column=5, value=self.start_date)
        ws_brief.cell(row=3, column=5, value=self.end_date)
        ws_brief.cell(row=4, column=5, value=self.timespan)
        queue_sheets = wb.get_sheet_names()
        queue_sheets.pop(0)
        for num, queue in enumerate(queue_sheets):
            for col, queue_info in enumerate(['B3', 'C5', 'F5', 'F6', 'D3'], start=1):
                ws_brief.cell(row=num + 11, column=col, value='={0}!{1}'.format(queue, queue_info))
        wb.save(self.tmpbillPath_for_admin_and_or_user)


    def excel_ready_to_zip(self):
        """
        将账单文件压缩为 zip 文件并存档
        在项目中不要随便切换目录，可能有多个用户要操作，会导致目录混乱
        """
        if self.day != 1:
            # 为 user 存档
            if not os.path.exists(os.path.dirname(self.billPath_for_user)):
                os.makedirs(os.path.dirname(self.billPath_for_user))
            shutil.move(self.tmpbillPath_for_admin_and_or_user, self.billPath_for_user)
        elif self.day == 1 and self.timespan == calendar.monthrange(self.year, self.month)[1]:  # 管理员查询日 判断 timespan 是否是当月天数
            if not os.path.exists(os.path.dirname(self.billPath_for_admin)):
                os.makedirs(os.path.dirname(self.billPath_for_admin))
            shutil.copyfile(self.tmpbillPath_for_admin_and_or_user, self.billPath_for_admin)
            if not os.path.exists(os.path.dirname(self.billPath_for_user)):
                os.makedirs(os.path.dirname(self.billPath_for_user))
            shutil.copyfile(self.tmpbillPath_for_admin_and_or_user, self.billPath_for_user)
            os.remove(self.tmpbillPath_for_admin_and_or_user)
        logging.info("用户 {0} 的计费账单 {1} 已生成".format(self.user_name, self.bill_name_for_user) + "\n")